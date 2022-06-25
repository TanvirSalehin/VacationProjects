#		Data Required

#	Teachers
Teachers_Start = "c3"
Teachers_End = "c6"

Total_Teachers = 4

#	Periods
Periods_Start = "d3"
Periods_End = "l6"




from openpyxl import load_workbook
def gen(n):
	for i in range(n):
		yield i
temp = gen(Total_Teachers)
def coords(ind2):
	temp = ind2//3
	match temp:
		case 0:
			day = "2"
		case 1:
			day = "3"
		case 2:
			day = "4"
		case 3:
			day = "5"
		case 4:
			day = "6"
		case 5:
			day = "7"

	temp2 = ind2 % 3
	match temp2:
		case 0:
			period = "b"
		case 1:
			period = "c"
		case 2:
			period = "d"
		case 2:
			period = "e"
		case 2:
			period = "f"
		case 2:
			period = "g"
		case 2:
			period = "h"
	return day, period

book = load_workbook("ROUTINE.xlsx")
Main = book.worksheets[0]
Main.title = "Main"
rows = Main.rows
days = [cell.value for cell in next(rows)]
periods = [cell.value for cell in next(rows)]
Teachers = []
for row in Main[Teachers_Start:Teachers_End]:
	for cell in row:
		Teachers.append(cell.value)
classes = {}
for teacher ,row in zip(Teachers, Main[Periods_Start:Periods_End]):
	classes_row = []
	for cell in row:
		classes_row.append(cell.value)
	classes[Teachers[next(temp)]] = classes_row


A8 = book.create_sheet("8A")
B8 = book.create_sheet("8B")
A9 = book.create_sheet("9A")
B9 = book.create_sheet("9B")
A10 = book.create_sheet("10A")
B10 = book.create_sheet("10B")
A11 = book.create_sheet("11A")
B11 = book.create_sheet("11B")
A12 = book.create_sheet("12A")
B12 = book.create_sheet("12B")

wsheets = [A8, A9, B8, B9, A10, B10, A11, B11, A12, B12]
Fnames = ["8A", "9A", "8B", "9B", "10A", "10B", "11A", "11B", "12A", "12B"]
for Form ,ws in zip(Fnames ,wsheets):
	ws["a2"] = 'Sunday'
	ws["a3"] = 'Monday'
	ws["a4"] = 'Tuesday'
	ws["a5"] = 'Wednesday'
	ws["a6"] = 'Thursday'
	
	ws["b1"] = 1
	ws["c1"] = 2
	ws["d1"] = 3
	ws["e1"] = 4
	ws["f1"] = 5
	ws["g1"] = 6
	ws["h1"] = 7

	print(Form)
	for index1 ,teacher in enumerate(classes):
		for index2, form in enumerate(classes[teacher]):
			if form == Form:
				print(index1, index2)
				x, y = coords(index2)
				ws[y+x] = Teachers[index1]


book.save("ROUTINE2.xlsx")
